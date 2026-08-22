#!/usr/bin/env python3
"""
Goettl LV — Warehouse & Fleet Movement Daily — data build (engine v7, 2026-08-22).

Routes input .xlsx files by HEADER SIGNATURE (filenames don't matter):
  * Movement  — header contains 'Transaction Type'   (Inventory - Profit Planner; MTD and/or single-day, any mix)
  * Snapshot  — header contains 'Quantity on Hand'   (Aggregate Inventory Stock Report; exactly the 2 newest by Filters date)
  * Jobs      — header contains 'Job #'              (Job Costing w Project info; any mix, deduped by Job #)
Optional: po_items.json in the input dir — Databricks fact_purchase_order_item pull (see WAREHOUSE_SOURCE_MAP.md §6).

Usage:
  python3 build_data.py --input <dir> [--template TEMPLATE.html --out live.html] [--built YYYY-MM-DD]
Emits data.json in CWD and prints a census. Exit 1 on a hard input failure.
All rules per WAREHOUSE_SOURCE_MAP.md — do not re-derive them here.
"""
import argparse, json, glob, os, re, sys, datetime, collections
import openpyxl

TAX = 0.08375                                   # Clark County — embedded in ST movement totals
TAXED = {'Receipt', 'Return', 'Bill', 'Purchase Order'}
CAGES = {'Vegas HVAC Warehouse', 'Vegas Plumbing Warehouse'}   # consumable rule: purchases = consumed on receipt
JH = 'GOETTL WAREHOUSE JOB HOLDING'
ACCT_WH = [('Vegas HVAC Warehouse', 'HVAC Warehouse (consumables cage)'),
           ('Vegas Safety Stock', 'Vegas Safety Stock'),
           ('Open Box Warehouse HVAC', 'Open Box HVAC'),
           ('Warranty Units HVAC', 'Warranty Units HVAC'),
           ('Vegas Plumbing Warehouse', 'Plumbing Warehouse (consumables cage)'),
           ('Plumbing Equipment', 'Plumbing Equipment')]
EQ_WH = ['Vegas Safety Stock', 'Open Box Warehouse HVAC', 'Warranty Units HVAC', 'Plumbing Equipment']

def is_truck(l): return l.startswith('LAS-') or l.startswith('Default Truck')
def fv(x):
    try: return float(x or 0)
    except (TypeError, ValueError): return 0.0
def sv(x):
    x = '' if x is None else str(x)
    return '' if x == 'N/A' else x.strip()

def sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c) if c is not None else '' for c in next(it, [])]
    return hdr, it, wb

def filters_date(wb):
    if 'Filters' not in wb.sheetnames: return None
    for r in wb['Filters'].iter_rows(values_only=True):
        cells = [str(c) for c in r if c is not None]
        if len(cells) >= 2 and cells[0].strip().lower() in ('date', 'date range'):
            m = re.search(r'(\d{2})/(\d{2})/(\d{2,4})\s*$', cells[1].strip())
            if m:
                mm, dd, yy = int(m[1]), int(m[2]), int(m[3])
                return datetime.date(yy + 2000 if yy < 100 else yy, mm, dd)
    return None

def eclass(name, desc=''):
    t = (name + ' ' + desc).upper()
    if t.strip() == 'EQUIPMENT': return 'Unspecified'
    if re.search(r'\b905\b.*\bAG\b|AIR GAP FAUCET', t): return 'WP Faucet (AG)'          # excluded from units tie per Stephen 2026-08-04
    if re.search(r'MINI ?SPLIT|DUCTLESS|^ASUM|\bASUM', t): return 'Mini-Split'
    if re.search(r'PACKAGE|PKG ?GAS|PKGGAS|[HM][- ]?PKGD|\bPKG\b|^GPG|\bGPG', t): return 'Package Unit'
    if re.search(r'AIR CONDITIONER|CONDENSER|HEAT PUMP|\bCOND/|\bCOND\b', t): return 'Condenser / HP'
    if re.search(r'^GLXS|^GLXT|^GLZS|^GSX|^GSZ|^ASX|^GZV|\bGLXS|\bGLXT|\bGLZS|\bGSX\w|\bASXC|\bGZV', t): return 'Condenser / HP'
    if re.search(r'FURNACE|GAS FURN|^GR9|^GRVT|^GM9|^GMVC|^ML180|^ML193|^ML296|^SL280|^EL296|\bGR9|\bGRVT|\bGMVC', t): return 'Furnace'
    if re.search(r'\bCOIL\b|^CAPT|^CHPT|^DP\d|^CP\d|^CK\d|\bCAPT|\bCHPT', t): return 'Coil'
    if re.search(r'AIR HANDLER|MULTIPOS|^AMST|^AVPTC|^ARUF|^AHVE|^LGM|\bAMST|\bAVPTC|\bAHVE|\bLGM', t): return 'Air Handler'
    if re.search(r'WATER HEATER|TANKLESS|\b\d+ GAL\b', t): return 'Water Heater'
    if re.search(r'SOFTEN|WATER CONDITIONER|REVERSE OSMOSIS|WATER MAKER|REFINER|SUPREME|CERTIFLOW|CARTRIDGE TANK|HALO|\bRO\b|5 STAR', t): return 'Water Purity'
    return 'Other Equipment'

# ---------- BILL OF MATERIALS (v4) ----------
# Maps a sold invoice TASK name to the component units it must consume. Confirmed by Stephen:
#   Split Gas   = condenser + coil + furnace (3)   Split Heat Pump = HP condenser + air handler (2)
#   Package gas / package HP = 1 package unit      Water heater (tank OR tankless) = 1, simple 1:1
#   Water purity: standalone softener/refiner = 1; refiner + softener = 2 ("most are 2 items")
# '1 Star IAQ Package' is deliberately NOT a Package Unit (IAQ package, would inflate by 31/90d).
BOM_SKIP = re.compile(r'REBUILD|FAN MOTOR|CONTROL BOARD|COMPRESSOR|DE-ICE|EASY START|SOFT START|ANODE'
                      r'|FLUSH|TEST & LIGHT|DRAIN VALVE|SMART SADIE|ZONE SYSTEM|REPIPE|DIAGNOSTIC'
                      r'|REMOVE |CLIENT SUPPLIED|WARR|CREDIT|DISCOUNT|MEMBERSH|REBATE|FINANC|LABOR'
                      r'|CLEAN|SEAL|DUCT|MAINT|TUNE|INSPECT|PERMIT|CRANE|DISPOS|CARTRIDGE'
                      r'|FILTER AND MEMBRANE|ELECTRICAL DISCONNECT|NIPPLE|HEX |\bIAQ\b|BIOCIDE|AEROSEAL')

def bom(name):
    t = (name or '').upper()
    if not t or BOM_SKIP.search(t): return None
    if re.search(r'SPLIT GAS', t):                                   return {'Condenser / HP': 1, 'Coil': 1, 'Furnace': 1}
    if re.search(r'SPLIT HEAT ?PUMP', t):                            return {'Condenser / HP': 1, 'Air Handler': 1}
    if re.search(r'PACKAGE (GAS|HEAT ?PUMP)', t):                    return {'Package Unit': 1}
    if re.search(r'(AC|HP) CONDENSER', t):                           return {'Condenser / HP': 1}
    if re.search(r'\bFURNACE\b', t):                                 return {'Furnace': 1}
    if re.search(r'CASED COIL|PLENUM COIL', t):                      return {'Coil': 1}
    if re.search(r'AIR HANDLER', t):                                 return {'Air Handler': 1}
    if re.search(r'WATER HEATER|TANKLESS SYSTEM SWAP', t):           return {'Water Heater': 1}
    if re.search(r'REFINER AND WATER SOFTENER|REFINER, WATER SOFTENER', t): return {'Water Purity': 2}
    if re.search(r'SOFTENING SYSTEM|\bREFINER\b|SUPREME|REVERSE OSMOSIS', t): return {'Water Purity': 1}
    return None

def install_class(jt):
    if re.match(r'^H-Install (Split|Package)', jt): return 'hvac_full'
    if jt == 'H-Install Partial': return 'hvac_partial'
    if jt == 'P-Install Water Purity': return 'wp'
    if jt == 'P-Install Water Heater': return 'wh'
    return None


# ============================== FLEET (v5, 2026-08-12) ==============================
# Inlined so the frozen artifact set stays exactly template.html + build_data.py + manifest.
# Rules live in WAREHOUSE_SOURCE_MAP v5 section 9.
# accounting's locked template names + par per truck (from month_end_inventory_v2 tmpl_defaults)
TMAP = {'SVC-H': 'HVAC Service', 'CFT-H': 'HVAC Craftsman', 'INS-H': 'HVAC Install',
        'SVC-P': 'Plumbing Service', 'INS-P': 'Plumbing Install', 'WP-P': 'Water Purity'}
PAR = {'HVAC Service': 9079.23, 'HVAC Craftsman': 8003.96, 'HVAC Install': 598.29,
       'Plumbing Service': 4432.87, 'Plumbing Install': 1392.88, 'Water Purity': 1384.74}
# accounting's fleet groups — the roll-up the month-end export reports on
GROUPS = [('hvac_service', 'HVAC Service (Craftsman and Service)', ['HVAC Craftsman', 'HVAC Service'], 'HVAC'),
          ('hvac_install', 'HVAC Install (Install only)', ['HVAC Install'], 'HVAC'),
          ('plumbing', 'Plumbing (All plumbing templates)', ['Plumbing Service', 'Plumbing Install', 'Water Purity'], 'Plumbing')]
UNTEMPLATED = 'Unassigned / Other'
RETIRED = 'Retired / transferred'
PCT_TRIP, DOLLAR_TRIP = 0.15, 500.0

# The location naming convention doubles as a STATUS field (discovered 2026-08-12). The
# template-code slot carries RET / RNR (retired) and XFR (transferred out of branch), and the
# driver slot carries RET / RETIRED / XFR on the same vehicles. Every one is $0 — stripped before
# retirement. These are roster history, NOT operational trucks: they are reported in their own
# class and excluded from fleet value, par and anomaly tracking.
STATUS_CODES = {'RET': 'retired', 'RNR': 'retired', 'XFR': 'transferred'}
STATUS_DRIVERS = {'RET', 'RETIRED', 'XFR', 'TRANSFERRED'}

LOC = re.compile(r'^LAS-([A-Za-z0-9]+)-([A-Z]{2,4}-[A-Z])_(\d+)\s*(.*)$')


def is_fleet(loc):
    return loc.startswith('LAS-') or loc.startswith('Default Truck')


def parse_loc(loc):
    """-> dict(veh, code, badge, driver, tmpl, status) for any fleet location."""
    m = LOC.match(loc)
    if m:
        code = m.group(2)
        drv = (m.group(4) or '').strip()
        status = STATUS_CODES.get(code.rsplit('-', 1)[0], '')
        if drv.upper() in STATUS_DRIVERS:
            status = status or ('transferred' if drv.upper().startswith('XFR') else 'retired')
            drv = ''
        if drv.upper() in ('N/A', 'NA', ''):
            drv = ''
        return {'veh': m.group(1), 'code': code, 'badge': m.group(3), 'driver': drv,
                'status': status,
                'tmpl': RETIRED if status else TMAP.get(code, UNTEMPLATED)}
    drv = ''
    if loc.startswith('Default Truck - '):
        drv = loc[len('Default Truck - '):].strip()
    return {'veh': '', 'code': 'DEFAULT', 'badge': '', 'driver': drv, 'status': '',
            'tmpl': UNTEMPLATED}


def vkey(loc):
    """Stable fleet identity. Vehicle number for LAS-* (survives driver renames); full string for
    Default Truck rows, which carry no vehicle number."""
    m = LOC.match(loc)
    return m.group(1) if m else loc


def load_roster(path):
    """Truck Template Validator — the ROSTER OF RECORD (ServiceTitan).
    Authoritative for which trucks exist, their Inventory Template and their assigned Technician.
    The stock snapshot only lists locations that currently hold item rows, so a truck with no
    inventory is invisible there — the roster is what makes those visible."""
    hdr, it, _ = sheet(path)
    ix = {h: i for i, h in enumerate(hdr)}
    g = lambda r, c: sv(r[ix[c]]) if c in ix and ix[c] < len(r) else ''
    out = {}
    for r in it:
        nm = g(r, 'Truck Name')
        if not nm: continue
        out[vkey(nm)] = {'loc': nm, 'tech': g(r, 'Technician'),
                         'tmpl': g(r, 'Inventory Template') or UNTEMPLATED,
                         'truckid': g(r, 'TruckId')}
    return out


def fleet_from_snap(snap):
    """snap: {location: {item_code: {name, ty, q, v, mx, uc}}} -> {location: rolled dict}"""
    out = {}
    for loc, items in snap.items():
        if not is_fleet(loc):
            continue
        info = parse_loc(loc)
        val = round(sum(i['v'] for i in items.values()), 2)
        tgt = round(sum(i.get('mx', 0) * i.get('uc', 0) for i in items.values()), 2)
        # Quantity on Hand exists from 2026-08-12 (Stephen added the column), so unit counts are
        # real. Items with no pricebook cost still carry quantity but contribute $0 of value —
        # counted separately so a value-only read never hides physical stock.
        qty = round(sum(i.get('q', 0) for i in items.values()), 1)
        qty0 = round(sum(i.get('q', 0) for i in items.values() if not i.get('uc')), 1)
        blind = sum(1 for i in items.values() if not i.get('uc') and i.get('q'))
        out[vkey(loc)] = dict(info, loc=loc, key=vkey(loc), val=val, tgt=tgt, lines=len(items),
                              qty=qty, qty_nocost=qty0, qty_blind=blind,
                              par=PAR.get(info['tmpl'], 0.0))
    return out


def build_fleet(cur_snap, cur_date, prev_snap=None, prev_date=None, roster=None, roster_date=None,
                roster_prev=None):
    cur = fleet_from_snap(cur_snap)
    prev = fleet_from_snap(prev_snap) if prev_snap else {}

    # ---- roster overlay (ServiceTitan Truck Template Validator) --------------------------------
    # The roster wins on TEMPLATE and TECHNICIAN because it states them as explicit fields; parsing
    # them out of the location string was only ever a workaround for not having this file. Trucks in
    # the roster with no snapshot rows are ADDED at $0 so a truck holding nothing is still visible.
    if roster:
        for k, rr in roster.items():
            t = cur.get(k)
            if t is None:
                info = parse_loc(rr['loc'])
                t = cur[k] = dict(info, loc=rr['loc'], key=k, val=0.0, tgt=0.0, lines=0,
                                  qty=0.0, qty_nocost=0.0, qty_blind=0, par=0.0)
            if rr['tmpl'] and not t['status']:
                t['tmpl'] = rr['tmpl']
                t['par'] = PAR.get(rr['tmpl'], 0.0)
            if rr['tech']:
                t['driver'] = rr['tech']
            t['truckid'] = rr.get('truckid', '')
            t['in_roster'] = True
        for k, t in cur.items():
            t.setdefault('in_roster', False)

    # --- classify: retired / inactive / active ------------------------------------------------
    # Stephen 2026-08-12: a truck with NO driver assigned AND NO inventory is INACTIVE. Value and
    # movement are tracked for ACTIVE trucks only. Retired/transferred is its own class ahead of
    # that test — a retired truck is roster history, not an idle asset to go chase.
    for t in cur.values():
        t['cls'] = ('retired' if t['status']
                    else 'inactive' if (not t['driver'] and abs(t['val']) < 0.005)
                    else 'active')
        t['active'] = t['cls'] == 'active'
    active = {k: v for k, v in cur.items() if v['active']}
    inactive = sorted((v for v in cur.values() if v['cls'] == 'inactive'),
                      key=lambda x: (x['tmpl'], x['veh']))
    retired = sorted((v for v in cur.values() if v['cls'] == 'retired'),
                     key=lambda x: (x['status'], x['veh']))

    span = None
    if prev_date and cur_date:
        span = (datetime.date.fromisoformat(cur_date) - datetime.date.fromisoformat(prev_date)).days

    # --- roster changes -----------------------------------------------------------------------
    # Match on VEHICLE NUMBER, not the full location string: a driver change rewrites the string,
    # so string-keying reports one truck as both "added" and "removed" (it did — 578137, 578189 and
    # A94362 each showed up on both lists). Vehicle number is the stable identity.
    cur_by_veh = {v['veh']: v for v in cur.values() if v['veh']}
    prev_by_veh = {v['veh']: v for v in prev.values() if v['veh']}
    cur_nov = {v['loc']: v for v in cur.values() if not v['veh']}      # Default Truck rows
    prev_nov = {v['loc']: v for v in prev.values() if not v['veh']}

    # Roster changes are a ROSTER-vs-ROSTER question. Diffing the roster against snapshot-parsed
    # names manufactures false positives: the roster states the current driver while the prior
    # snapshot's location string states the driver as of that day, so every driver change since the
    # roster was pulled reads as a "reassignment", and every roster truck holding no stock reads as
    # "added". Measured 2026-08-12: that produced 7 false adds and 15 false reassignments. With only
    # one roster on hand, report NOTHING here and say why.
    roster_diff = bool(roster and roster_prev)
    added, removed, reassigned = [], [], []
    if roster and not roster_prev:
        pass
    elif roster_diff:
        for k, rr in roster.items():
            if k not in roster_prev:
                t = cur.get(k, {})
                added.append(dict(t or {}, loc=rr['loc'], tmpl=rr['tmpl'], driver=rr['tech'],
                                  val=(t or {}).get('val', 0.0), change='added'))
        for k, pr in roster_prev.items():
            if k not in roster:
                removed.append({'loc': pr['loc'], 'veh': k, 'tmpl': pr['tmpl'], 'driver': pr['tech'],
                                'status': '', 'val': (prev.get(k) or {}).get('val', 0.0),
                                'change': 'removed'})
            else:
                rr = roster[k]
                if pr['tech'] != rr['tech'] or pr['tmpl'] != rr['tmpl']:
                    t = cur.get(k, {})
                    reassigned.append({'veh': k, 'loc': rr['loc'], 'tmpl': rr['tmpl'],
                                       'tmpl_from': pr['tmpl'], 'status': '',
                                       'from': pr['tech'] or '(unassigned)',
                                       'to': rr['tech'] or '(unassigned)',
                                       'val': t.get('val', 0.0),
                                       'prev': (prev.get(k) or {}).get('val', 0.0)})
    elif prev:
        for veh, v in cur_by_veh.items():
            if veh not in prev_by_veh:
                added.append(dict(v, change='added'))
        for veh, p in prev_by_veh.items():
            c = cur_by_veh.get(veh)
            if not c:
                removed.append(dict(p, change='removed'))
            elif p['driver'] != c['driver'] or p['tmpl'] != c['tmpl']:
                reassigned.append({'veh': veh, 'loc': c['loc'], 'tmpl': c['tmpl'],
                                   'tmpl_from': p['tmpl'], 'status': c['status'],
                                   'from': p['driver'] or '(unassigned)',
                                   'to': c['driver'] or '(unassigned)',
                                   'val': c['val'], 'prev': p['val']})
        for loc, v in cur_nov.items():
            if loc not in prev_nov:
                added.append(dict(v, change='added'))
        for loc, p in prev_nov.items():
            if loc not in cur_nov:
                removed.append(dict(p, change='removed'))

    # --- per-truck rows, active only ----------------------------------------------------------
    rows = []
    for loc, t in active.items():
        p = prev.get(loc)
        p0 = p['val'] if p else None
        dv = round(t['val'] - p0, 2) if p0 is not None else None
        fill = round(100.0 * t['val'] / t['par'], 1) if t['par'] else None
        flags = []
        if t['val'] < -0.005:
            flags.append('negative')
        if dv is not None:
            if t['par'] and abs(dv) >= PCT_TRIP * t['par']:
                flags.append('pct')
            if abs(dv) >= DOLLAR_TRIP:
                flags.append('dollar')
        if not t['driver']:
            flags.append('nodriver')
        rows.append(dict(t, prev=p0, dv=dv, fill=fill, flags=flags,
                         anom=bool({'negative', 'pct', 'dollar'} & set(flags))))
    rows.sort(key=lambda r: -abs(r['dv'] if r['dv'] is not None else 0))

    # --- template roll-up (active trucks only) ------------------------------------------------
    tmpl = collections.OrderedDict()
    for r in rows:
        d = tmpl.setdefault(r['tmpl'], {'tmpl': r['tmpl'], 'n': 0, 'val': 0.0, 'prev': 0.0,
                                        'tgt': 0.0, 'par_each': PAR.get(r['tmpl'], 0.0),
                                        'par': 0.0, 'anom': 0, 'nodriver': 0})
        d['n'] += 1
        d['val'] += r['val']
        d['prev'] += (r['prev'] or 0.0)
        d['tgt'] += r['tgt']
        d['par'] += r['par']
        d['anom'] += 1 if r['anom'] else 0
        d['nodriver'] += 1 if not r['driver'] else 0
    for d in tmpl.values():
        for k in ('val', 'prev', 'tgt', 'par'):
            d[k] = round(d[k], 2)
        d['dv'] = round(d['val'] - d['prev'], 2) if prev else None
        d['fill'] = round(100.0 * d['val'] / d['par'], 1) if d['par'] else None

    order = [t for _, _, ts, _ in GROUPS for t in ts] + [UNTEMPLATED]
    tmpl_rows = sorted(tmpl.values(), key=lambda d: order.index(d['tmpl']) if d['tmpl'] in order else 99)

    # --- fleet groups -> trade ----------------------------------------------------------------
    grows = []
    for key, label, members, trade in GROUPS:
        mem = [d for d in tmpl_rows if d['tmpl'] in members]
        grows.append({'key': key, 'label': label, 'trade': trade, 'templates': members,
                      'n': sum(d['n'] for d in mem),
                      'val': round(sum(d['val'] for d in mem), 2),
                      'prev': round(sum(d['prev'] for d in mem), 2),
                      'par': round(sum(d['par'] for d in mem), 2),
                      'anom': sum(d['anom'] for d in mem)})
    un = [d for d in tmpl_rows if d['tmpl'] == UNTEMPLATED]
    if un:
        grows.append({'key': 'untemplated', 'label': 'Untemplated / Default trucks', 'trade': 'Unassigned',
                      'templates': [UNTEMPLATED], 'n': sum(d['n'] for d in un),
                      'val': round(sum(d['val'] for d in un), 2),
                      'prev': round(sum(d['prev'] for d in un), 2), 'par': 0.0,
                      'anom': sum(d['anom'] for d in un)})
    for g in grows:
        g['dv'] = round(g['val'] - g['prev'], 2) if prev else None

    trade = collections.OrderedDict()
    for g in grows:
        d = trade.setdefault(g['trade'], {'trade': g['trade'], 'n': 0, 'val': 0.0, 'prev': 0.0, 'par': 0.0, 'anom': 0})
        for k in ('n', 'anom'):
            d[k] += g[k]
        for k in ('val', 'prev', 'par'):
            d[k] = round(d[k] + g[k], 2)
    for d in trade.values():
        d['dv'] = round(d['val'] - d['prev'], 2) if prev else None

    tot = {'n': len(rows), 'inactive': len(inactive), 'retired': len(retired),
           'val': round(sum(r['val'] for r in rows), 2),
           'prev': round(sum((r['prev'] or 0.0) for r in rows), 2),
           'par': round(sum(r['par'] for r in rows), 2),
           'tgt': round(sum(r['tgt'] for r in rows), 2),
           'qty': round(sum(r['qty'] for r in rows), 1),
           'qty_nocost': round(sum(r['qty_nocost'] for r in rows), 1),
           'anom': sum(1 for r in rows if r['anom']),
           'nodriver': sum(1 for r in rows if not r['driver']),
           'qty_blind': sum(r['qty_blind'] for r in rows)}
    tot['dv'] = round(tot['val'] - tot['prev'], 2) if prev else None
    tot['fill'] = round(100.0 * tot['val'] / tot['par'], 1) if tot['par'] else None

    roster_note = None
    if roster and roster_date and roster_date != cur_date:
        roster_note = (f'Roster is dated {roster_date}, snapshot {cur_date} — template and technician '
                       f'are from the roster, so a change made between those dates reads as current.')
    return {'roster_date': roster_date, 'roster_n': len(roster or {}), 'roster_note': roster_note,
            'roster_diff': roster_diff,
            'cur_date': cur_date, 'prev_date': prev_date, 'span': span,
            'has_prev': bool(prev), 'tot': tot, 'tmpl': tmpl_rows, 'groups': grows,
            'trades': list(trade.values()), 'trucks': rows, 'inactive': inactive,
            'retired': retired,
            'added': added, 'removed': removed, 'reassigned': reassigned,
            'thresholds': {'pct': PCT_TRIP, 'dollar': DOLLAR_TRIP},
            'par_table': PAR, 'tmap': TMAP}

# ============================ end FLEET block ============================

# ServiceTitan 'Equipment Type' -> this engine's class vocabulary. Used only to resolve units the
# name classifier could not place (see the Equipment branch below). 'Configurable' and blank stay
# unresolved on purpose — they are a pricebook gap, and hiding them would hide the fix.
ST_EQTYPE = {
    'Straight Cool Condensers R32': 'Condenser / HP',
    'Heat Pump Condensers R32':     'Condenser / HP',
    'Split Air Conditioner':        'Condenser / HP',
    'Furnace R32':                  'Furnace',
    'Plenum Coil R32':              'Coil',
    'Plenum Coil':                  'Coil',
    'Evaporator Coil R32':          'Coil',
    'Air Handler R32':              'Air Handler',
    'Packaged Heat Pump':           'Package Unit',
    'Water Heater':                 'Water Heater',
}

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True)
    ap.add_argument('--template'); ap.add_argument('--out')
    ap.add_argument('--built', default=None)
    ap.add_argument('--history', default=None, help='history.json path — day-keyed accumulation, 92-day retention')
    ap.add_argument('--notes', default=None, help='notes.json — approved explanations exported from the page, baked into DATA.notes')
    a = ap.parse_args()

    # Router (v4): header signature first, Filters-tab report title only to break exact ties.
    # NOTE 2026-08-05: 'Jobs trailing 90 days' and 'Real Discounts' ship IDENTICAL headers AND
    # identical payloads — header signature alone cannot separate them, so invoice-item files are
    # content-deduped below. Units Installed also matches 'Item GL Group Name'+'Item Quantity',
    # so it is claimed FIRST on its unique 'Completion Date' column.
    movement, snaps, jobsfiles, soldfiles, taskfiles, rosterfiles = [], [], [], [], [], []
    for p in sorted(glob.glob(os.path.join(a.input, '*.xlsx'))):
        if os.path.basename(p).startswith('~$'): continue
        try: hdr, _, _ = sheet(p)
        except Exception: continue
        # Roster claimed FIRST on its own unique pair — it is the fleet roster of record (§9).
        if 'Truck Name' in hdr and 'Inventory Template' in hdr: rosterfiles.append(p)
        elif 'Transaction Type' in hdr: movement.append(p)
        elif 'Quantity on Hand' in hdr: snaps.append(p)
        elif 'Completion Date' in hdr and 'Item GL Group Name' in hdr: soldfiles.append(p)
        elif 'Invoice Number' in hdr and 'Item Type' in hdr: taskfiles.append(p)
        elif 'Item GL Group Name' in hdr and 'Item Quantity' in hdr: soldfiles.append(p)
        elif 'Job #' in hdr: jobsfiles.append(p)
    if not movement and not os.path.exists(a.history or ''): sys.exit('FATAL: no movement export and no history')

    # ---------- movement ledger (dedupe by (type,doc,date,loc); tax-stripped; Bills dropped) ----------
    seen, ledger, bills = set(), [], []
    jobstat = {}          # v4: job_number -> live ServiceTitan Job Status, straight off the movement export.
                          # 2,407 jobs / 90 days vs 490 in the MTD Job Costing file. This is THE fix for the
                          # false "open — job not completed" labels (job 550193 reads Completed here).
    for p in movement:
        hdr, it, _ = sheet(p)
        ix = {h: i for i, h in enumerate(hdr)}
        g = lambda r, c: sv(r[ix[c]]) if c in ix else ''
        for r in it:
            tt = g(r, 'Transaction Type')
            jn_s, js_s = g(r, 'Job Number'), g(r, 'Job Status')
            if jn_s and js_s: jobstat[jn_s] = js_s      # captured BEFORE the Bill skip — Bills carry status too
            if tt == 'Bill':
                # v5.5 Leakage L3 (map §12L): Bills stay OUT of the ledger (accounting mirror of
                # Receipts, §3) — captured here for the bill tripwires ONLY. Doc grammar is locked:
                # PO '<base>' · Receipt '<base>-R<n>' · Bill '<base>-B<n>'.
                _brv = r[ix['Date Received']] if 'Date Received' in ix and ix['Date Received'] < len(r) else None
                bills.append({'doc': g(r, 'Transaction Number'), 'job': jn_s,
                              'd': _brv.strftime('%Y-%m-%d') if hasattr(_brv, 'strftime') else sv(_brv)[:10],
                              'v': round(fv(r[ix['Total']]) / (1 + TAX), 2)})
                continue
            if not tt: continue
            # PHYSICAL-DATE RULE (2026-08-13). 'Date' / 'Date Created' is when a human keyed the
            # transaction, not when material moved: on the 08-11 export every row carries Date
            # Created = 08-11 while Date Received splits 08-11/08-12, Date Approved reaches back to
            # 07-10 and Date Sent to 08-04. Dating on the keystroke made a catch-up entry day look
            # like a huge movement day and broke the whole point of a day-specific view.
            # Receipts/Bills -> Date Received. Transfers -> Date Picked, else Date Sent.
            # Returns / Adjustments / POs have no physical stamp, so they keep 'Date'.
            def _iso(v):
                return v.strftime('%Y-%m-%d') if hasattr(v, 'strftime') else (sv(v)[:10] if sv(v) not in ('', 'N/A') else '')
            def _col(c):
                return _iso(r[ix[c]]) if c in ix and ix[c] < len(r) else ''
            keyed = _col('Date') or _col('Date Created')
            if tt in ('Receipt', 'Bill'):
                iso = _col('Date Received') or keyed
            elif 'Transfer' in tt:
                iso = _col('Date Picked') or _col('Date Sent') or keyed
            else:
                iso = keyed
            if not iso: iso = keyed
            key = (tt, g(r, 'Transaction Number'), iso, g(r, 'Inventory Location'), g(r, 'Transfer From'), g(r, 'Transfer To'))
            if key in seen: continue
            seen.add(key)
            gross = fv(r[ix['Total']])
            row = {'t': tt, 'd': iso, 'loc': g(r, 'Inventory Location'), 'from': g(r, 'Transfer From'),
                   'to': g(r, 'Transfer To'), 'v': round(gross / (1 + TAX), 2) if tt in TAXED else round(gross, 2),
                   'doc': g(r, 'Transaction Number'), 'vendor': g(r, 'Vendor Name'), 'tech': g(r, 'Technician Name'),
                   'pot': g(r, 'Purchase Order Type'), 'at': g(r, 'Adjustment Type'), 'rt': g(r, 'Return Type'),
                   'job': g(r, 'Job Number'), 'st': g(r, 'Transaction Status'),
                   'dk': keyed if keyed and keyed != iso else '',
                   'qty': fv(g(r, 'Quantity on Transaction')) if 'Quantity on Transaction' in ix else 0,
                   'stid': g(r, 'Job ID') if 'Job ID' in ix else '',
                   'appr': g(r, 'Approval Status')}
            ledger.append({k: v for k, v in row.items() if v or k in ('t', 'd', 'v')})
    days = sorted({r['d'] for r in ledger})

    # ---------- jobs / installs ----------
    jmap = {}
    for p in jobsfiles:
        hdr, it, _ = sheet(p)
        ix = {h: i for i, h in enumerate(hdr)}
        for r in it:
            jn = sv(r[ix['Job #']])
            if not jn: continue
            end = r[ix['End of Working Time']]
            # 'Job ID' is ServiceTitan's INTERNAL id and the only thing /#/Job/Index/<id> accepts —
            # the job NUMBER will not resolve. Job Costing is MTD, so links only reach back to the
            # 1st until Job ID is added to the trailing-90 report.
            jmap[jn] = {'jt': sv(r[ix['Job Type']]), 'tech': sv(r[ix['Assigned Technicians']]),
                        'stid': sv(r[ix['Job ID']]) if 'Job ID' in ix else '',
                        'end': end.strftime('%Y-%m-%d') if hasattr(end, 'strftime') else ''}
    inst_by_day = collections.defaultdict(lambda: {'hvac_full': 0, 'hvac_partial': 0, 'wp': 0, 'wh': 0})
    for j in jmap.values():
        c = install_class(j['jt'])
        if c and j['end']: inst_by_day[j['end']][c] += 1

    # ---------- invoice items + BILL OF MATERIALS (v4, 'Jobs trailing 90 days' feed) ----------
    # Invoice TASKS are the demand signal: a sold "Split Gas System" means 1 condenser + 1 coil + 1
    # furnace were used, with certainty. Recipes confirmed by Stephen 2026-08-04 / 2026-08-05.
    # Content-dedupe: 'Real Discounts' ships the same payload as 'Jobs trailing 90 days'.
    inv_eq, bom_job, task_rows, seen_inv = [], collections.defaultdict(collections.Counter), [], set()
    qty_viol = []
    for p in taskfiles:
        hdr, it, _ = sheet(p)
        ix = {h: i for i, h in enumerate(hdr)}
        def gi(r, c):
            i = ix.get(c)
            return sv(r[i]) if i is not None and i < len(r) else ''
        for r in it:
            inv = gi(r, 'Invoice Number')
            if not inv: continue
            nm, ity, jn = gi(r, 'Item Name'), gi(r, 'Item Type'), gi(r, 'Job Number')
            q = fv(gi(r, 'Item Quantity')) or 1.0
            k = (inv, nm, gi(r, 'Item Price'), ity, gi(r, 'Item Cost'), gi(r, 'Item Quantity'))
            if k in seen_inv: continue          # identical duplicate feed (Real Discounts)
            seen_inv.add(k)
            d = r[ix['Invoice Date']] if 'Invoice Date' in ix and ix['Invoice Date'] < len(r) else None
            iso = d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else sv(d)[:10]
            if ity == 'Service':
                b = bom(nm)
                if b:
                    # Guard: Item Quantity on service tasks is not always a whole count — the feed
                    # carries 0.27 / 0.5 / 1.5 on labor lines and NEGATIVE values (e.g. 'Need
                    # Drywaller' at -1.8). A sold system is a whole unit; clamp to a positive integer
                    # so a stray fraction can never inflate or subtract expected component units.
                    # Sign-aware HALF-UP rounding. Python's round() is banker's rounding, so
                    # round(0.5) == 0 and a half-unit task would vanish. Negatives are credit /
                    # reversal lines (a re-invoiced job carries +1 and -1 of the same task) and MUST
                    # subtract so the pair nets to zero — treating -1 as +1 double-counted the system.
                    qn = (1 if q >= 0 else -1) * int(abs(q) + 0.5) if q else 1
                    if qn == 0: qn = 1 if q > 0 else 0
                    if q != qn or q < 0:
                        # Stephen 2026-08-05: a partial quantity is fine on labor, but AC units, water
                        # heaters and water purity systems are ALWAYS whole 1:1 physical equipment.
                        # A fraction on one of those breaks the rule — surface the job, never silently clamp.
                        qty_viol.append({'job': jn, 'task': nm[:70], 'q': q, 'used': qn,
                                         'kind': 'reversal' if q < 0 else 'partial',
                                         'cls': ', '.join(sorted(b)), 'd': iso,
                                         'tech': gi(r, 'Assigned Technicians')[:24],
                                         'jt': gi(r, 'Job Type')})
                    for cls, n in b.items(): bom_job[jn][cls] += n * qn
                    q = qn
                    task_rows.append({'d': iso, 'job': jn, 'task': nm[:70], 'q': q,
                                      'jt': gi(r, 'Job Type'), 'bom': b,
                                      'tech': gi(r, 'Assigned Technicians')[:24]})
            elif ity == 'Equipment':
                c = eclass(nm)
                if c == 'WP Faucet (AG)': continue        # excluded per Stephen
                # 2026-08-13: the trailing-90 invoice feed now carries ServiceTitan's own curated
                # 'Equipment Type', plus 'Serial Number' and 'Installed On' per unit. Per the
                # trust-the-curated-field rule the ST class is used to RESOLVE units the name-based
                # classifier could not place — it does NOT override a class that already resolved,
                # so the locked variance baselines cannot shift underneath us.
                et = gi(r, 'Equipment Type')
                if c in ('Unspecified', 'Other Equipment', '') and et:
                    c = ST_EQTYPE.get(et.strip(), c)
                inst = gi(r, 'Installed On')[:10]
                inv_eq.append({'d': iso, 'job': jn, 'name': nm[:46], 'cls': c, 'q': q,
                               'jt': gi(r, 'Job Type'),
                               'et': et, 'ser': gi(r, 'Serial Number')[:32],
                               'id': inst, 'inv': gi(r, 'Is Inventory')})
            if jn and gi(r, 'Job Type'):
                jmap.setdefault(jn, {'jt': gi(r, 'Job Type'),
                                     'tech': gi(r, 'Assigned Technicians'), 'stid': '', 'end': ''})
                # 'Job ID' landed on the trailing-90 feed 2026-08-13 (verified 226/226 against the
                # movement export). Job Costing is MTD only, so this is what gives every job on the
                # page a working ServiceTitan link across the full 90 days.
                if not jmap[jn].get('stid'):
                    jmap[jn]['stid'] = gi(r, 'Job ID')

    # ---------- sold units (Units Installed report) ----------
    # Day key = Completion Date (fallback Invoice Date). Files may overlap days (trailing-14 window):
    # per-day newest-file-wins, files ranked by (has Job Number column, max data date).
    def load_sold(p):
        hdr, it, _ = sheet(p)
        ix = {h: i for i, h in enumerate(hdr)}
        dcol = 'Completion Date' if 'Completion Date' in ix else 'Invoice Date'
        jcol = next((h for h in hdr if h.strip().lower() in ('job', 'job #', 'job number')), None)
        byd = {}
        for r in it:
            nm = sv(r[ix['Item Name']])
            if not nm: continue
            d = r[ix[dcol]]
            iso = d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else sv(d)[:10]
            byd.setdefault(iso, []).append({'d': iso, 'name': nm[:46], 'code': sv(r[ix['Item Code']]),
                                            'cls': eclass(nm), 'q': fv(r[ix['Item Quantity']]),
                                            'jt': sv(r[ix['Job Type']]), 'job': sv(r[ix[jcol]]) if jcol else ''})
        return byd, bool(jcol)
    sold_days = {}
    ranked = []
    for p in soldfiles:
        byd, hasjob = load_sold(p)
        ranked.append((hasjob, max(byd) if byd else '', byd))
    for hasjob, mx, byd in sorted(ranked, key=lambda x: (x[0], x[1])):   # best file applied last → wins overlapping days
        sold_days.update(byd)
    sold = [s for d in sorted(sold_days) for s in sold_days[d] if s['cls'] != 'WP Faucet (AG)']   # faucets excluded per Stephen 2026-08-04

    # ---------- snapshots (2 newest by Filters date) ----------
    warnings, month, skus, cls_roll, tie = [], None, [], {}, []
    # v5: also carries Max/Min Quantity and Item Unit Cost — the truck REPLENISHMENT TEMPLATE, which
    # is what makes the Fleet tab's par math possible. 'Quantity on Hand' was added to the daily
    # subscription 2026-08-12, so the daily aggregate export now routes here on the same signature
    # as the month-end one and no router change is needed.
    def loadsnap(p):
        hdr, it, _ = sheet(p)
        ix = {h: i for i, h in enumerate(hdr)}
        g = lambda r, c: r[ix[c]] if c in ix and ix[c] < len(r) else None
        out = {}
        for r in it:
            loc = sv(g(r, 'Inventory Location'))
            if not loc: continue
            d = out.setdefault(loc, {})
            c = sv(g(r, 'Item Code'))
            uc = fv(g(r, 'Item Unit Cost'))
            itm = d.setdefault(c, {'name': sv(g(r, 'Item Name'))[:46], 'desc': sv(g(r, 'Item Description'))[:90],
                                   'ty': sv(g(r, 'Item Type')), 'q': 0.0, 'v': 0.0,
                                   'mx': 0.0, 'mn': 0.0, 'uc': uc})
            itm['q'] += fv(g(r, 'Quantity on Hand')); itm['v'] += fv(g(r, 'Total Item Cost'))
            itm['mx'] += fv(g(r, 'Max Quantity')); itm['mn'] += fv(g(r, 'Min Quantity'))
            if uc: itm['uc'] = uc
        return out

    snap0 = snap1 = None; d0 = d1 = None
    dated = sorted(((filters_date(sheet(p)[2]), p) for p in snaps), key=lambda x: (x[0] or datetime.date.min))
    if len(dated) >= 2:
        (d0, p0), (d1, p1) = dated[-2], dated[-1]
        if len(dated) > 2: warnings.append(f'{len(dated)} snapshots found — using {d0} and {d1}.')
        snap0, snap1 = loadsnap(p0), loadsnap(p1)
    elif len(dated) == 1:
        # One snapshot still drives the Fleet tab at levels-only. Do NOT set snap0 — the month panel
        # and equipment stock math genuinely need two and must stay omitted rather than compare a
        # snapshot against nothing.
        d1, p1 = dated[0]
        snap1 = loadsnap(p1)
        warnings.append('Only 1 stock snapshot — Fleet reports levels with no movement; Month Tie-Out omitted.')
    else:
        warnings.append('No stock snapshots — Fleet tab and Month Tie-Out omitted.')

    # ---------- optional Databricks PO items ----------
    # Generic catch-all PO lines (SKU literally named 'Materials' / 'Equipment' / 'Misc'). A unit
    # bought on one of these is invisible to every downstream system — the SKU name carries no class,
    # even though the buyer typed a real description on the line. Verified on job 553837: the plenum
    # coil was bought as 'Materials' $608.00 on the same PO as the condenser and furnace. These are
    # surfaced against the job so a 'unit not found' reads as a pricebook fix, not a mystery.
    generic_po = collections.defaultdict(list)
    gj = os.path.join(a.input, 'generic_po.json')
    if os.path.exists(gj):
        gp = json.load(open(gj))
        gi = {c: i for i, c in enumerate(gp['cols'])}
        # Genie widens an IN list to ILIKE, so real named SKUs ('Equipment Rentals', 'E-Lite
        # Equipment Pad', 'Field Po Materials') come back too. Those DO identify what was bought —
        # keep only the true catch-alls, whose names carry no class at all.
        CATCHALL = {'materials', 'material', 'equipment', 'misc', 'miscellaneous', 'parts'}
        for r in gp['rows']:
            job = str(r[gi['job_number']] or '')
            if not job: continue
            if str(r[gi['item_name']] or '').strip().lower() not in CATCHALL: continue
            generic_po[job].append({'name': str(r[gi['item_name']] or ''), 'ty': str(r[gi['item_type']] or ''),
                                    'q': fv(r[gi['quantity']]), 'v': round(fv(r[gi['total']]), 2),
                                    'po': str(r[gi['purchase_order_number']] or ''),
                                    'd': str(r[gi['purchase_order_date']] or '')[:10]})

    eq_po = []
    pj = os.path.join(a.input, 'po_items.json')
    if os.path.exists(pj):
        po = json.load(open(pj))
        ci = {c: i for i, c in enumerate(po['cols'])}
        for r in po['rows']:
            name = r[ci['item_name']] or ''; job = r[ci['job_number']]
            eq_po.append({'d': r[ci['purchase_order_date']], 'name': name[:46], 'cls': eclass(name),
                          'q': fv(r[ci['quantity']]), 'v': round(fv(r[ci['total']]), 2), 'job': str(job or ''),
                          'vendor': (r[ci['primary_vendor_name']] or '')[:30], 'po': str(r[ci['purchase_order_number']] or ''),
                          'dest': 'Direct to job' if job else 'Stock replenishment', 'st': 'stock', 'j': None})
    else:
        warnings.append('po_items.json missing — equipment ledger empty this run.')

    # ---------- month panel + equipment stock math (snapshot window) ----------
    if snap0:
        w0, w1 = (d0 or datetime.date.min).isoformat(), (d1 or datetime.date.max).isoformat()
        inwin = lambda d: w0 < d <= w1
        mv = collections.defaultdict(lambda: collections.defaultdict(float))
        mvn = collections.defaultdict(lambda: collections.defaultdict(int))
        for r in ledger:
            if not inwin(r['d']): continue
            loc = r.get('loc', '')
            if r['t'] == 'Receipt': mv[loc]['rec'] += r['v']; mvn[loc]['nrec'] += 1
            elif r['t'] == 'Return': mv[loc]['ret'] += r['v']; mvn[loc]['nret'] += 1
            elif r['t'] == 'Adjustment': mv[loc]['adj'] += r['v']; mvn[loc]['nadj'] += 1
            elif 'Transfer' in r['t']: mvn[r.get('to', '')]['xin'] += 1; mvn[r.get('from', '')]['xout'] += 1
        whval = lambda sn, loc: round(sum(i['v'] for i in sn.get(loc, {}).values()), 2)
        def top_items(loc, n=6):
            diffs = []
            for c in set(snap0.get(loc, {})) | set(snap1.get(loc, {})):
                i0, i1 = snap0.get(loc, {}).get(c), snap1.get(loc, {}).get(c)
                dv = (i1['v'] if i1 else 0) - (i0['v'] if i0 else 0)
                if abs(dv) < 0.005: continue
                ref = i1 or i0
                diffs.append({'name': ref['name'][:44], 'dq': (i1['q'] if i1 else 0) - (i0['q'] if i0 else 0), 'dv': round(dv, 2)})
            return sorted(diffs, key=lambda x: -abs(x['dv']))[:n]
        month_wh = []
        for loc, label in ACCT_WH:
            o, c = whval(snap0, loc), whval(snap1, loc)
            m, n = mv[loc], mvn[loc]
            valued = round(m['rec'] - m['ret'] + m['adj'], 2)
            month_wh.append({'loc': loc, 'label': label, 'open': o, 'close': c, 'var': round(c - o, 2),
                             'rec': round(m['rec'], 2), 'ret': round(m['ret'], 2), 'adj': round(m['adj'], 2),
                             'nrec': n['nrec'], 'nret': n['nret'], 'nadj': n['nadj'], 'xin': n['xin'], 'xout': n['xout'],
                             'resid': round(c - o - valued, 2), 'cage': loc in CAGES, 'items': top_items(loc)})
        jh = {'open': whval(snap0, JH), 'close': whval(snap1, JH), 'rec': round(mv[JH]['rec'], 2),
              'nrec': mvn[JH]['nrec'], 'adj': round(mv[JH]['adj'], 2), 'items': top_items(JH, 8)}
        odd = [{'loc': l, 'open': whval(snap0, l), 'close': whval(snap1, l), 'var': round(whval(snap1, l) - whval(snap0, l), 2)}
               for l in set(list(snap0) + list(snap1))
               if not is_truck(l) and l not in dict(ACCT_WH) and l != JH and (abs(whval(snap0, l)) > 0.005 or abs(whval(snap1, l)) > 0.005)]
        trucks = {'open': round(sum(whval(snap0, l) for l in snap0 if is_truck(l)), 2),
                  'close': round(sum(whval(snap1, l) for l in snap1 if is_truck(l)), 2),
                  'adj': round(sum(r['v'] for r in ledger if r['t'] == 'Adjustment' and is_truck(r.get('loc', '')) and inwin(r['d'])), 2)}
        month = {'wh': month_wh, 'jh': jh, 'odd': odd, 'trucks': trucks}
        for wh in EQ_WH:
            for c in set(snap0.get(wh, {})) | set(snap1.get(wh, {})):
                i0, i1 = snap0.get(wh, {}).get(c), snap1.get(wh, {}).get(c)
                ref = i1 or i0
                if ref['ty'] != 'Equipment': continue
                q0, v0 = (i0['q'], i0['v']) if i0 else (0, 0); q1, v1 = (i1['q'], i1['v']) if i1 else (0, 0)
                skus.append({'wh': wh, 'name': ref['name'], 'cls': eclass(ref['name'], ref['desc']),
                             'q0': q0, 'q1': q1, 'dq': q1 - q0, 'v0': round(v0, 2), 'v1': round(v1, 2), 'dv': round(v1 - v0, 2)})
        for s in skus:
            r = cls_roll.setdefault(s['cls'], {'q0': 0, 'q1': 0, 'dq': 0, 'dv': 0.0})
            r['q0'] += s['q0']; r['q1'] += s['q1']; r['dq'] += s['dq']; r['dv'] = round(r['dv'] + s['dv'], 2)
        # RECEIPT-DATE RULE (v5.5, map §12L): inbound units are dated by the ServiceTitan RECEIPT
        # ('Date Received' of the '<PO base>-R<n>' row), never by purchase_order_date — a PO dated a
        # week before delivery would land outside a daily window and make receipt day read as phantom
        # gains (proof: Certiflow PO 646571076, PO-dated 08-06, received 08-13, 42 units). A PO with
        # no receipt row has not arrived and contributes nothing. Multi-receipt POs: earliest receipt.
        rec_d = {}
        for r in ledger:
            if r['t'] != 'Receipt' or not r.get('doc'): continue
            _b = re.sub(r'-R\d+$', '', r['doc'])
            if _b and r.get('d') and (_b not in rec_d or r['d'] < rec_d[_b]): rec_d[_b] = r['d']
        inb, direct = collections.defaultdict(float), collections.defaultdict(float)
        for e in eq_po:
            rd = rec_d.get(e.get('po', ''))
            if rd is None or not inwin(rd): continue
            (inb if e['dest'] == 'Stock replenishment' else direct)[e['cls']] += e['q']
        for cls, r in cls_roll.items():
            so = r['q0'] + inb[cls] - r['q1']
            tie.append({'cls': cls, 'q0': r['q0'], 'in': inb[cls], 'q1': r['q1'], 'stock_out': round(so, 1),
                        'direct': direct[cls], 'total_to_jobs': round(so + direct[cls], 1)})
        inst_tot = collections.Counter()
        for j in jmap.values():
            c = install_class(j['jt'])
            if c and j['end'] and inwin(j['end']): inst_tot[c] += 1
    else:
        inst_tot = collections.Counter()

    # ---------- fleet (v5) ----------
    # Snapshot-scoped, NOT window-scoped: it reports fleet state on the snapshot date and the change
    # since the previous snapshot. Built off the same two snapshots the month panel uses.
    roster, roster_date = None, None
    if rosterfiles:
        rp = sorted(rosterfiles)[-1]
        roster = load_roster(rp)
        rd = filters_date(sheet(rp)[2])
        roster_date = rd.isoformat() if rd else None
        if not roster_date:
            # Validator carries no Filters date — fall back to the file mtime so staleness is visible.
            roster_date = datetime.date.fromtimestamp(os.path.getmtime(rp)).isoformat()
    else:
        warnings.append('Truck Template Validator missing — fleet template/technician fall back to '
                        'parsing the location string, and trucks holding no stock are invisible.')

    # Prior roster: same report, an older dated copy. Only with two can roster changes be reported.
    roster_prev = None
    if len(rosterfiles) >= 2:
        roster_prev = load_roster(sorted(rosterfiles)[-2])

    fleet = None
    if snap1:
        fleet = build_fleet(snap1, d1.isoformat() if d1 else '',
                            snap0, d0.isoformat() if (snap0 and d0) else None,
                            roster, roster_date, roster_prev)
        ft = fleet['tot']
        if ft['anom']:
            warnings.append(f"Fleet: {ft['anom']} truck(s) past threshold "
                            f"({int(PCT_TRIP*100)}% of par or ${int(DOLLAR_TRIP)}) or holding negative value.")
        neg = [t['loc'] for t in fleet['trucks'] if t['val'] < -0.005]
        if neg:
            warnings.append('Fleet NEGATIVE on-hand value (impossible — bad count or mis-post): ' + ', '.join(neg[:6]))
        bad_removals = [t['veh'] or t['loc'] for t in fleet['removed'] if abs(t['val']) > 0.005]
        if bad_removals:
            warnings.append('Fleet vehicles left the roster still carrying value: ' + ', '.join(bad_removals[:6]))

    # ---------- fleet template composition (v7, map §14.8) ----------
    # The live ServiceTitan min/max template per accounting template, taken as the MODE across that
    # template's ACTIVE trucks (items with Max Quantity > 0 only). This is what powers the Month End
    # 4-cause fleet variance drill: (1) vendor cost change on a template item, (2) item added to the
    # template, (3) item removed, (4) truck added/removed. Saved into each month close by the page's
    # "Close the month" button so closed months can be diffed item-by-item, never recomputed.
    if fleet and snap1:
        _by_t = collections.defaultdict(lambda: collections.defaultdict(list))   # tmpl -> code -> [(mx,uc,name)]
        _tn = collections.Counter()
        for t in fleet['trucks']:                                # ACTIVE trucks only
            # Bucket by the RAW location-string template (TMAP on the code), NOT the roster-overlaid
            # t['tmpl']: the min/max rows on a truck are what ServiceTitan stamped for the location's
            # own template, so a stale roster retag must not leak one truck's items into another
            # template's composition (measured 2026-08-22: 59 phantom 'added' items on HVAC Service).
            _rawt = TMAP.get(t.get('code', ''), None)
            if not _rawt: continue
            _tn[_rawt] += 1
            for c, i in (snap1.get(t['loc'], {}) or {}).items():
                if (i.get('mx') or 0) > 0:
                    _by_t[_rawt][c].append((i['mx'], i.get('uc') or 0.0, i['name']))
        tcomp = {}
        for _tm, _items in _by_t.items():
            comp, drift = {}, 0
            for c, obs in _items.items():
                mxm = collections.Counter(x[0] for x in obs).most_common(1)[0][0]
                ucm = collections.Counter(x[1] for x in obs).most_common(1)[0][0]
                if any(x[0] != mxm or abs(x[1] - ucm) > 0.005 for x in obs): drift += 1
                comp[c] = {'name': obs[0][2][:44], 'mx': round(mxm, 1), 'uc': round(ucm, 2),
                           'on': len(obs)}                       # trucks carrying the item row
            tcomp[_tm] = {'n': _tn[_tm], 'items': comp, 'drift_items': drift,
                          'tgt_each': round(sum(v['mx'] * v['uc'] for v in comp.values()), 2)}
        fleet['tcomp'] = tcomp

    # ---------- fleet roster EVENTS — raw snapshot-pair diff (v7, map §14.8) ----------
    # A second, independent channel from the roster-gated added/removed lists above: both sides are
    # RAW parsed snapshot locations (apples to apples, vehicle-number keyed), so it is immune to the
    # stale-roster overlay problem and safe to accumulate nightly. Powers vehicle-level naming in the
    # Month End fleet variance drill. Fleet-tab roster reporting is UNCHANGED.
    fev = None
    if snap0 and snap1 and fleet:
        _c, _p = fleet_from_snap(snap1), fleet_from_snap(snap0)
        _cb = {v['veh']: v for v in _c.values() if v['veh']}
        _pb = {v['veh']: v for v in _p.values() if v['veh']}
        _cn = {v['loc']: v for v in _c.values() if not v['veh']}
        _pn = {v['loc']: v for v in _p.values() if not v['veh']}
        fev = {'a': [], 'r': [], 'x': []}
        for k, v in _cb.items():
            p = _pb.get(k)
            if p is None:
                fev['a'].append({'veh': k, 'tmpl': v['tmpl'], 'driver': v['driver'], 'val': round(v['val'], 2)})
            elif p['driver'] != v['driver'] or p['tmpl'] != v['tmpl']:
                fev['x'].append({'veh': k, 'tmpl': v['tmpl'], 'tmpl_from': p['tmpl'],
                                 'from': p['driver'] or '(unassigned)', 'to': v['driver'] or '(unassigned)'})
        for k, p in _pb.items():
            if k not in _cb:
                fev['r'].append({'veh': k, 'tmpl': p['tmpl'], 'driver': p['driver'], 'val': round(p['val'], 2)})
        for k, v in _cn.items():
            if k not in _pn:
                fev['a'].append({'veh': k[:40], 'tmpl': v['tmpl'], 'driver': v['driver'], 'val': round(v['val'], 2)})
        for k, p in _pn.items():
            if k not in _cn:
                fev['r'].append({'veh': k[:40], 'tmpl': p['tmpl'], 'driver': p['driver'], 'val': round(p['val'], 2)})
        fleet['events'] = {d1.isoformat(): fev}

    installs = [{'d': j['end'], 'job': jn, 'jt': j['jt'], 'cls': install_class(j['jt']), 'tech': j['tech'][:24]}
                for jn, j in jmap.items() if install_class(j['jt']) and j['end']]

    # ---------- history: day-keyed upsert, 92-day retention ----------
    if a.history:
        H = json.load(open(a.history)) if os.path.exists(a.history) else {'ledger': {}, 'sold': {}, 'installs': {}, 'po': {}}
        def upsert(store, rows, datekey='d'):
            byd = {}
            for r in rows: byd.setdefault(r[datekey], []).append(r)
            for d, rs in byd.items(): store[d] = rs           # replace whole day (self-heal)
        upsert(H['ledger'], ledger); upsert(H['sold'], sold)
        upsert(H['installs'], installs); upsert(H['po'], eq_po)
        # v7: accumulate the raw snapshot-pair fleet events by snapshot date, same 92-day retention.
        if fev is not None and d1:
            H.setdefault('fleet_events', {})[d1.isoformat()] = fev
        cutoff = (datetime.date.fromisoformat(max(list(H['ledger']) + ['1970-01-01'])) - datetime.timedelta(days=92)).isoformat()
        for k in ('ledger', 'sold', 'installs', 'po'):
            H[k] = {d: v for d, v in H[k].items() if d >= cutoff}
        if 'fleet_events' in H:
            H['fleet_events'] = {d: v for d, v in H['fleet_events'].items() if d >= cutoff}
            if fleet is not None:
                fleet['events'] = H['fleet_events']
        json.dump(H, open(a.history, 'w'), default=str)
        ledger = [r for d in sorted(H['ledger']) for r in H['ledger'][d]]
        sold = [r for d in sorted(H['sold']) for r in H['sold'][d]]
        installs = [r for d in sorted(H['installs']) for r in H['installs'][d]]
        eq_po = [r for d in sorted(H['po']) for r in H['po'][d]]
        days = sorted(set(H['ledger']) | set(H['sold']))   # installs excluded — End of Working Time can stray outside the feed window

    # ---------- v4: job completion resolution + ledger buckets + three-tier variance ----------
    # Runs AFTER the history merge so every historical PO row is re-stated with today's knowledge.
    # Completion date priority: Units Installed Completion Date > Job Costing 'End of Working Time'
    # > invoice date on the job's equipment lines. Status priority: movement Job Status (authoritative,
    # live, 90 days) > presence of a completion date.
    sold_end, inv_end = {}, {}
    for s in sold:
        if s.get('job'): sold_end.setdefault(s['job'], s['d'])
    for e in inv_eq:
        if e.get('job') and e.get('d'): inv_end.setdefault(e['job'], e['d'])
    DONE = {'Completed'}
    DEAD = {'Canceled', 'Cancelled'}
    OPEN = {'Scheduled', 'InProgress', 'Hold'}

    def resolve(job):
        """-> (state, end_date_or_None). state in matched|open|canceled|nofeed."""
        if not job: return 'stock', None
        st = jobstat.get(job, '')
        end = sold_end.get(job) or (jmap.get(job, {}) or {}).get('end') or inv_end.get(job)
        if st in DEAD: return 'canceled', end
        if st in DONE: return 'matched', end
        if st in OPEN: return 'open', end
        if end:        return 'matched', end          # completion evidence with no status row
        return 'nofeed', None                          # NEVER call this "open" — that was the bug

    restated = collections.Counter()
    for e in eq_po:
        state, end = resolve(e.get('job', ''))
        j0 = e.get('st')
        e['st'] = state
        e['j'] = None if state in ('stock', 'nofeed') else {
            'end': end or '', 'jt': (jmap.get(e['job'], {}) or {}).get('jt', ''),
            'tech': ((jmap.get(e['job'], {}) or {}).get('tech') or '')[:24]}
        if j0 != state: restated[f'{j0}->{state}'] += 1
    if restated:
        warnings.append('PO status restated (v4 job-status fix): ' + ', '.join(f'{k} {v}' for k, v in restated.most_common()))

    # --- ledger buckets: consumed on job / returned / missing / received -------------------------
    # Stephen 2026-08-05: anything that left for a job — bought for it OR pulled to it, staged or
    # installed — is ONE bucket. Missing = vendor receipt exists, destination does not.
    def bucket(e):
        if e['st'] in ('matched', 'open', 'canceled'): return 'consumed'
        if e['st'] == 'nofeed': return 'missing'
        return 'received'
    for e in eq_po:
        e['bk'] = bucket(e)
        e['flight'] = (e['st'] == 'open')            # left the warehouse, job not closed yet
    ret_rows = [r for r in ledger if r['t'] == 'Return'
                or (r['t'] == 'Adjustment' and r.get('at') in ('ReturnToVendor', 'ReturnToWarehouse'))]
    buckets = {}
    for key in ('consumed', 'returned', 'missing', 'received'):
        # SCOPE HONESTY: consumed / missing / received are UNIT-level (equipment PO line items, so
        # "count by type" is real). Returned can only be transaction-level — the movement export has
        # no Item Name / Item Quantity, so returns are documents-and-dollars grouped by location, and
        # the bucket is tagged scope='txn' so the page never presents a document count as a unit count.
        txn = (key == 'returned')
        src = ret_rows if txn else [e for e in eq_po if e['bk'] == key]
        by = collections.defaultdict(lambda: {'n': 0.0, 'v': 0.0})
        for e in src:
            grp = (e.get('loc') or 'Unknown location') if txn else e.get('cls', 'Other Equipment')
            by[grp]['n'] += 1 if txn else (e.get('q', 1) or 1)
            by[grp]['v'] += abs(e.get('v', 0) or 0) if txn else (e.get('v', 0) or 0)
        buckets[key] = {'cls': {k: {'n': round(v['n'], 1), 'v': round(v['v'], 2)} for k, v in by.items()},
                        'n': round(sum(v['n'] for v in by.values()), 1),
                        'v': round(sum(v['v'] for v in by.values()), 2),
                        'rows': len(src), 'scope': 'txn' if txn else 'unit',
                        'items': ([{'d': r['d'], 'loc': r.get('loc', ''), 'v': abs(r.get('v', 0)),
                                    'doc': r.get('doc', ''), 'vendor': r.get('vendor', '') or r.get('tech', ''),
                                    'rt': r.get('rt', '') or r.get('at', '')} for r in src] if txn else
                                  [{'d': e['d'], 'name': e['name'], 'cls': e['cls'], 'q': e.get('q', 1),
                                    'v': e.get('v', 0), 'job': e.get('job', ''), 'vendor': e.get('vendor', ''),
                                    'po': e.get('po', ''), 'st': e['st'], 'flight': e.get('flight', False),
                                    'end': (e.get('j') or {}).get('end', '')} for e in src])}

    # --- three-tier variance: required (BOM) vs purchased-to-job vs invoiced --------------------
    po_job = collections.defaultdict(collections.Counter)
    for e in eq_po:
        if e.get('job'): po_job[e['job']][e['cls']] += e.get('q', 1) or 1
    inv_job = collections.defaultdict(collections.Counter)
    for e in inv_eq:
        if e.get('job'): inv_job[e['job']][e['cls']] += e.get('q', 1) or 1
    CLS = ['Condenser / HP', 'Coil', 'Furnace', 'Air Handler', 'Package Unit', 'Water Heater', 'Water Purity']

    # UNIT-RESCUE (v5.6, map §7a fourth pull — job 554442 lesson). Real units ride PO lines whose
    # pricebook SKU is NOT typed Equipment (a $1,104 water heater keyed on a Material-typed
    # "2 GAL EXP TNK" SKU), so the type-filtered po_items pull can't see them and the job reads
    # 'stock'/'short' falsely. Rescue lines are ONE-DIRECTIONAL evidence: they can SATISFY an unmet
    # requirement (clearing a false flag) but can NEVER create an over-pull or a notask — applied
    # per job/class only up to max(0, required − equipment-typed PO units). Engine-side gate:
    # eclass(name+description) must resolve to a tracked class, total >= $150, and the line must not
    # match the accessory skip list (motors, curbs, condensate pumps, pans, stands, cleaners...).
    RESCUE_SKIP = re.compile(r'MOTOR|CURB|CONDENSATE|CLEANER|\bPAN\b|STAND|BRACKET|FILTER|DRIER'
                             r'|CAPACITOR|CONTACTOR|THERMOSTAT|BREAKER|DISCONNECT|WHIP|\bPAD\b')
    rescue_job = collections.defaultdict(list)
    rjp = os.path.join(a.input, 'po_rescue.json')
    if os.path.exists(rjp):
        rp = json.load(open(rjp))
        ri = {c: i for i, c in enumerate(rp['cols'])}
        for r in rp['rows']:
            nm = str(r[ri['item_name']] or ''); ds = str(r[ri['item_description']] or '')
            job = str(r[ri['job_number']] or '')
            if not job or fv(r[ri['total']]) < 150: continue
            if RESCUE_SKIP.search((nm + ' ' + ds).upper()): continue
            c = eclass(nm, ds)
            if c not in CLS: continue
            rescue_job[job].append({'cls': c, 'q': fv(r[ri['quantity']]) or 1.0,
                                    'v': round(fv(r[ri['total']]), 2), 'name': nm[:46],
                                    'po': str(r[ri['purchase_order_number']] or ''),
                                    'd': str(r[ri['purchase_order_date']] or '')[:10],
                                    'ty': str(r[ri['item_type']] or '')})
    else:
        warnings.append('po_rescue.json missing — mis-typed-SKU unit rescue off; jobs whose unit rides a non-Equipment SKU may read stock/short falsely (see job 554442).')
    branch = []
    for c in CLS:
        req = sum(b[c] for b in bom_job.values())
        branch.append({'cls': c, 'req': round(req, 1),
                       'inv': round(sum(b[c] for b in inv_job.values()), 1),
                       'poj': round(sum(b[c] for b in po_job.values()), 1),
                       'pos': round(sum(e.get('q', 1) for e in eq_po if e['cls'] == c and not e.get('job')), 1)})
    jobvar = []
    for job in set(list(bom_job) + list(po_job) + list(inv_job)):
        state, end = resolve(job)
        rows = []
        for c in set(list(bom_job[job]) + list(po_job[job]) + list(inv_job[job])):
            req, pj, iv = bom_job[job][c], po_job[job][c], inv_job[job][c]
            if not (req or pj or iv): continue
            # v5.6 unit-rescue: mis-typed-SKU lines satisfy an unmet requirement, capped at the gap
            # so they can never push a class into over-pull.
            resq = 0.0
            if req > 0 and pj < req and job in rescue_job:
                avail = sum(x['q'] for x in rescue_job[job] if x['cls'] == c)
                resq = min(avail, req - pj)
            rows.append({'cls': c, 'req': round(req, 1), 'po': round(pj, 1), 'inv': round(iv, 1),
                         'resq': round(resq, 1),
                         'vmove': round(pj - req, 1), 'vinv': round(iv - req, 1)})
        if not rows: continue
        # Flag precedence. 'over' is Stephen's waste case and REQUIRES a known requirement:
        # required 1 condenser, 2 PO'd to the job -> +1. A job with units moved but NO recognised
        # system task (required 0) is NOT a over-pull — that is a task-classification / itemisation
        # problem and gets its own flag, otherwise it floods the red list with false positives.
        moved = sum(r['po'] for r in rows)
        anyreq = any(r['req'] for r in rows)
        # A required unit with no job PO and nothing itemized splits two ways, and conflating them
        # made a 177-job red list out of 24 real items. If the job carries a generic catch-all PO
        # line, that line is almost certainly the unit on the wrong SKU -> 'short', actionable.
        # If it carries nothing at all, the unit came off warehouse stock, which ServiceTitan does
        # not record per job -> 'stock', the DOCUMENTED BLIND SPOT, informational not red.
        nounit = any(r['req'] and not (r['po'] + r.get('resq', 0)) and not r['inv'] for r in rows)
        # 'short' USED to mean inv < req — but Goettl invoices FLAT RATE, so a sold system carries a
        # Service task and no Item Type='Equipment' row at all. That made inv==0 the normal case and
        # fired on 233 of 776 jobs, including job 553887 whose coil was correctly PO'd and visible.
        # Corrected 2026-08-12: only flag when a required class has NO evidence at all — nothing
        # PO'd to the job AND nothing itemized. If the unit was bought against the job, it is
        # accounted for; whether the invoice names it is a pricebook question, not a variance.
        flag = ('over'     if any(r['vmove'] > 0 and r['req'] for r in rows)
                else 'canceled' if state == 'canceled' and moved
                else 'notask'   if moved and not anyreq
                else 'short'    if (nounit and generic_po.get(job))
                else 'stock'    if nounit else 'ok')
        # Window anchor for the variance sections is the job's COMPLETION date — that is when the
        # sale happened and when units must reconcile. Filtering each side by its own date would split
        # pairs at a window edge (a PO dated 7/21 against a job that closed 7/22). Where completion is
        # unknown, fall back to the earliest PO date on the job so the row still lands in a window
        # instead of disappearing from every one of them; endEst marks it as estimated.
        anchor, est = end, False
        if not anchor:
            pds = [e['d'] for e in eq_po if e.get('job') == job and e.get('d')]
            anchor = min(pds) if pds else inv_end.get(job) or sold_end.get(job) or ''
            est = bool(anchor)
        jobvar.append({'job': job, 'st': state, 'end': end or '', 'anchor': anchor or '',
                       'endEst': est, 'flag': flag, 'rows': rows,
                       'gen': generic_po.get(job, []),
                       'resc': rescue_job.get(job, []),
                       'stid': (jmap.get(job, {}) or {}).get('stid', ''),
                       'jt': (jmap.get(job, {}) or {}).get('jt', ''),
                       'tech': ((jmap.get(job, {}) or {}).get('tech') or '')[:24]})
    _ORD = {'over': 0, 'canceled': 1, 'notask': 2, 'short': 3, 'stock': 4, 'ok': 5}
    jobvar.sort(key=lambda x: (_ORD.get(x['flag'], 9), x['anchor'] or '', x['job']))

    # ---------- LEAKAGE tab (v5.5, map §12L) ----------
    OB = 'Open Box Warehouse HVAC'
    _wv = lambda sn, loc: round(sum(i['v'] for i in (sn or {}).get(loc, {}).values()), 2)
    sd = None
    if snap1:
        ob_items = []
        if snap0:
            for c in set(snap0.get(OB, {})) | set(snap1.get(OB, {})):
                i0, i1 = snap0.get(OB, {}).get(c), snap1.get(OB, {}).get(c)
                dq = (i1['q'] if i1 else 0) - (i0['q'] if i0 else 0)
                dv = (i1['v'] if i1 else 0) - (i0['v'] if i0 else 0)
                if abs(dq) < 0.005 and abs(dv) < 0.005: continue
                ref = i1 or i0
                ob_items.append({'name': ref['name'], 'code': c, 'dq': round(dq, 1), 'dv': round(dv, 2),
                                 'q1': round(i1['q'], 1) if i1 else 0, 'v1': round(i1['v'], 2) if i1 else 0})
        _cur = snap1.get(OB, {})
        sd = {'d0': d0.isoformat() if (snap0 and d0) else None, 'd1': d1.isoformat() if d1 else None,
              'open': _wv(snap0, OB) if snap0 else None, 'close': _wv(snap1, OB),
              'units': round(sum(i['q'] for i in _cur.values()), 1),
              'nitems': sum(1 for i in _cur.values() if abs(i['q']) > 0.005),
              'items': sorted(ob_items, key=lambda x: -abs(x['dv']))}

    # L2 — trucks over the live ServiceTitan template (>110%, Stephen 2026-08-14). INS-H is grouped
    # as stale-template (§12L); untemplated actives listed as hygiene — a % rule cannot see them.
    trucks_over, untmpl = [], []
    if fleet and snap1:
        for t in fleet['trucks']:
            _tgt, _val = t.get('tgt') or 0.0, t.get('val') or 0.0
            if _tgt > 0 and _val > 1.10 * _tgt:
                _items = []
                for c, i in (snap1.get(t['loc'], {}) or {}).items():
                    _ov = (i['q'] - (i['mx'] or 0)) * (i['uc'] or 0)
                    if i['q'] > (i['mx'] or 0) and _ov > 0.005:
                        _items.append({'name': i['name'], 'q': round(i['q'], 1),
                                       'mx': round(i['mx'] or 0, 1), 'ov': round(_ov, 2)})
                _items.sort(key=lambda x: -x['ov'])
                trucks_over.append({'loc': t['loc'], 'veh': t.get('veh', ''), 'drv': t.get('driver') or '',
                                    'code': t.get('code', ''), 'tmpl': t.get('tmpl', ''),
                                    'val': round(_val, 2), 'tgt': round(_tgt, 2),
                                    'over': round(_val - _tgt, 2), 'pct': round(100.0 * _val / _tgt, 1),
                                    'stale': t.get('code') == 'INS-H', 'items': _items[:8]})
            elif _tgt <= 0 and _val > 0.005:
                untmpl.append({'loc': t['loc'], 'veh': t.get('veh', ''), 'drv': t.get('driver') or '',
                               'code': t.get('code', ''), 'val': round(_val, 2)})
        trucks_over.sort(key=lambda x: -x['over'])
        untmpl.sort(key=lambda x: -x['val'])

    # L3 bill tripwires — verified dormant 2026-08-14 (0 orphans / 0 mismatches on 5,873 bills).
    po_doc, rec_bases = {}, set()
    for r in ledger:
        if r['t'] == 'Purchase Order' and r.get('doc'): po_doc[r['doc']] = r
        elif r['t'] == 'Receipt' and r.get('doc'): rec_bases.add(re.sub(r'-R\d+$', '', r['doc']))
    orph, mism, seen_b = [], [], set()
    for bl in bills:
        if not bl.get('doc') or bl['doc'] in seen_b: continue
        seen_b.add(bl['doc'])
        _base = re.sub(r'-B\d+$', '', bl['doc'])
        _po = po_doc.get(_base)
        if _po is None and _base not in rec_bases:
            orph.append(bl)
        elif _po is not None and bool(_po.get('job')) != bool(bl.get('job')):
            mism.append({'doc': bl['doc'], 'pojob': _po.get('job', ''), 'billjob': bl.get('job', ''),
                         'd': bl.get('d', '')})
    bills_flags = {'n': len(seen_b), 'orphans': orph[:50], 'mismatch': mism[:50]}

    # L4 — Small Tools SKU lines (third pinned Genie pull, §7a). Vendor comes from the matching
    # ledger PO row — the SKU's primary vendor is just the pricebook default.
    tools = None
    stj = os.path.join(a.input, 'small_tools.json')
    if os.path.exists(stj):
        sp = json.load(open(stj))
        si = {c: i for i, c in enumerate(sp['cols'])}
        tools = []
        for r in sp['rows']:
            _pon = str(r[si['purchase_order_number']] or '')
            tools.append({'d': str(r[si['purchase_order_date']] or '')[:10], 'po': _pon,
                          'q': fv(r[si['quantity']]), 'v': round(fv(r[si['total']]), 2),
                          'job': str(r[si['job_number']] or ''),
                          'vendor': (po_doc.get(_pon, {}).get('vendor') or '')[:30]})
    else:
        warnings.append('small_tools.json missing — Leakage L4 runs on the ServiceTitan PO-type channel alone.')

    leak = {'sd': sd, 'trucks': {'over': trucks_over, 'untmpl': untmpl, 'thresh': 110},
            'bills': bills_flags, 'tools': tools}

    # ---------- month closes (v6) — prior-month CLOSING balances, saved not recomputed ----------
    # The Month End tab compares this month's running position against the two prior closes. Those
    # are read from month_closes.json and NEVER recalculated: once a month is closed and handed to
    # accounting, its number is frozen — recomputing it from a later snapshot would silently restate
    # a period accounting has already booked. Seeded 2026-08-19 from the 06/30 and 07/31 daily
    # aggregate snapshots through this engine's own loadsnap + build_fleet, so the prior columns and
    # the live column are produced by identical math. Each later month is appended by the page's
    # "Close the month" button.
    mclose = {}
    mcj = os.path.join(a.input, 'month_closes.json')
    if os.path.exists(mcj):
        try:
            mclose = json.load(open(mcj))
        except Exception:
            warnings.append('month_closes.json unreadable — Month End prior-month columns omitted.')
    else:
        warnings.append('month_closes.json missing — Month End tab runs current-month-only, no prior-month comparison.')

    notes = {}
    if a.notes and os.path.exists(a.notes):
        try: notes = json.load(open(a.notes))
        except Exception: warnings.append('notes.json unreadable — baked notes skipped')
    elif a.notes is None and os.path.exists(os.path.join(a.input, 'notes.json')):
        try: notes = json.load(open(os.path.join(a.input, 'notes.json')))
        except Exception: warnings.append('notes.json unreadable — baked notes skipped')

    data = {'built': a.built or datetime.date.today().isoformat(), 'days': days, 'ledger': ledger,
            'month': month, 'warnings': warnings, 'sold': sold, 'notes': notes,
            'equip': {'skus': sorted(skus, key=lambda x: -abs(x['dv'])), 'cls': cls_roll, 'po': eq_po,
                      'instDay': inst_by_day, 'instTot': dict(inst_tot), 'tie': tie,
                      'installs': sorted(installs, key=lambda x: x['d'])},
            'mclose': mclose,
            'bk': buckets, 'branch': branch, 'jobvar': jobvar, 'tasks': task_rows, 'qtyViol': qty_viol,
            'invEq': inv_eq, 'jobstat_n': len(jobstat), 'fleet': fleet, 'leak': leak,
            'stids': {k: v.get('stid', '') for k, v in jmap.items() if v.get('stid')},
            'stBase': 'https://goettl_lasvegas.eh.go.servicetitan.com/#/Job/Index/'}
    json.dump(data, open('data.json', 'w'), default=str)
    stc = collections.Counter(e['st'] for e in eq_po)
    print(json.dumps({'ok': True, 'ledger_rows': len(ledger), 'days': f'{days[0]}..{days[-1]}' if days else '—',
                      'movement_files': len(movement), 'snapshots': [str(d0), str(d1)] if snap0 else None,
                      'jobs': len(jmap), 'po_items': len(eq_po), 'sold_rows': len(sold), 'history': bool(a.history),
                      'jobstat_jobs': len(jobstat), 'po_status': dict(stc),
                      'task_rows': len(task_rows), 'inv_eq_rows': len(inv_eq),
                      'bom_jobs': len(bom_job), 'jobvar_rows': len(jobvar),
                      'flags': dict(collections.Counter(j['flag'] for j in jobvar)),
                      'qty_violations': len(qty_viol),
                      'leak': {'sd_items': len((sd or {}).get('items', [])),
                               'sd_close': (sd or {}).get('close'),
                               'trucks_over': len(trucks_over), 'untemplated': len(untmpl),
                               'bills_checked': bills_flags['n'],
                               'bill_orphans': len(bills_flags['orphans']),
                               'bill_job_mismatch': len(bills_flags['mismatch']),
                               'tool_lines': (len(tools) if tools is not None else None),
                               'rescue_lines': sum(len(v) for v in rescue_job.values()),
                               'rescue_jobs': len(rescue_job)},
                      'generic_po_jobs': len(generic_po),
                      'generic_po_value': round(sum(x['v'] for v in generic_po.values() for x in v), 2),
                      'stids': len({k: v for k, v in jmap.items() if v.get('stid')}),
                      'fleet': (None if not fleet else {
                          'snapshot': fleet['cur_date'], 'prior': fleet['prev_date'], 'span': fleet['span'],
                          'active': fleet['tot']['n'], 'inactive': fleet['tot']['inactive'],
                          'retired': fleet['tot']['retired'], 'value': fleet['tot']['val'],
                          'par': fleet['tot']['par'], 'delta': fleet['tot']['dv'],
                          'anomalies': fleet['tot']['anom'],
                          'tcomp_templates': len(fleet.get('tcomp', {})),
                          'roster': [len(fleet['added']), len(fleet['removed']), len(fleet['reassigned'])]}),
                      'buckets': {k: (v['n'], v['v']) for k, v in buckets.items()},
                      'branch': {b['cls']: (b['req'], b['inv'], b['poj'], b['pos']) for b in branch},
                      'warnings': warnings}, indent=1))

    if a.template and a.out:
        tpl = open(a.template, encoding='utf-8').read()
        assert '/*__DATA__*/null' in tpl, 'splice marker missing in template'
        open(a.out, 'w', encoding='utf-8').write(tpl.replace('/*__DATA__*/null', json.dumps(data, separators=(",", ":"), default=str)))
        print('spliced ->', a.out)

if __name__ == '__main__':
    main()
